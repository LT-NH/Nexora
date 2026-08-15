"""Excel/CSV export endpoints using openpyxl and the csv module.

Supports two output formats (``excel`` / ``csv``) selected via the
``format`` query parameter, optional date-range filtering
(``date_from`` / ``date_to``), and non-blocking file generation by
offloading synchronous ``openpyxl`` / ``csv`` work to a worker thread
via :func:`asyncio.to_thread`.
"""
import asyncio
import csv
import io
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_workspace
from app.db.session import get_session
from app.models import Order, Product

router = APIRouter(prefix="/workspaces/{workspace_slug}/export", tags=["export"])

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2560EB", end_color="2560EB", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")

PRODUCT_HEADERS = [
    "商品名称", "SKU", "分类", "价格", "库存", "低库存预警", "条码", "状态",
]
ORDER_HEADERS = [
    "订单号", "客户姓名", "金额", "状态", "物流单号", "承运商", "创建时间",
]

_VALID_FORMATS = ("excel", "csv")


def _parse_date_bounds(
    date_from: Optional[str], date_to: Optional[str]
) -> tuple[Optional[datetime], Optional[datetime]]:
    """Parse ``date_from`` / ``date_to`` strings into timezone-aware bounds.

    Returns a ``(dt_from, dt_to)`` tuple where either element may be
    ``None`` when the corresponding parameter is not supplied.

    Raises:
        HTTPException 422: When a supplied value is not ``YYYY-MM-DD``.
    """
    dt_from: Optional[datetime] = None
    dt_to: Optional[datetime] = None

    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d").replace(
                tzinfo=timezone.utc
            )
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Invalid date_from format. Use YYYY-MM-DD.",
            )

    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc
            )
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Invalid date_to format. Use YYYY-MM-DD.",
            )

    return dt_from, dt_to


def _product_row(p: Product) -> list:
    """Build a flat row list for a single product."""
    return [
        p.name,
        p.sku or "",
        getattr(p, "category", "") or "",
        float(p.price),
        getattr(p, "stock", 0),
        getattr(p, "low_stock_threshold", 10),
        getattr(p, "barcode", "") or "",
        p.status.value if hasattr(p.status, "value") else str(p.status),
    ]


def _order_row(o: Order) -> list:
    """Build a flat row list for a single order."""
    return [
        getattr(o, "order_number", o.id[:8]),
        getattr(o, "customer_name", "") or "",
        float(o.total),
        o.status.value if hasattr(o.status, "value") else str(o.status),
        getattr(o, "tracking_number", "") or "",
        getattr(o, "carrier", "") or "",
        str(o.created_at)[:10],
    ]


def _write_excel(
    rows: list[list], headers: list[str], sheet_title: str
) -> io.BytesIO:
    """Build an Excel workbook in memory (synchronous, CPU-bound)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_csv(rows: list[list], headers: list[str]) -> io.BytesIO:
    """Build a CSV file in memory (synchronous).

    Uses ``utf-8-sig`` (BOM) encoding so the file opens correctly in
    Excel with Chinese characters preserved.
    """
    text_buf = io.StringIO()
    writer = csv.writer(text_buf)
    writer.writerow(headers)
    writer.writerows(rows)
    output = io.BytesIO()
    output.write(text_buf.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return output


async def _build_export_response(
    rows: list[list],
    headers: list[str],
    sheet_title: str,
    filename_base: str,
    fmt: str,
) -> StreamingResponse:
    """Build a :class:`StreamingResponse` for the requested format.

    The synchronous file generation (``openpyxl`` or ``csv``) is executed
    in a worker thread via :func:`asyncio.to_thread` to avoid blocking the
    event loop for large exports. Because the work is already offloaded
    from the event loop, these endpoints intentionally do NOT go through
    the background task queue — an export is a user-requested, synchronous
    operation whose result is streamed back, not a fire-and-forget job.
    """
    if fmt == "csv":
        output = await asyncio.to_thread(_write_csv, rows, headers)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename_base}.csv",
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

    # Default: excel
    output = await asyncio.to_thread(_write_excel, rows, headers, sheet_title)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename_base}.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.get("/products")
async def export_products(
    workspace_slug: str,
    format: str = Query("excel", description="Output format: excel or csv"),
    date_from: Optional[str] = Query(
        None, description="Filter records created on or after this date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(
        None, description="Filter records created on or before this date (YYYY-MM-DD)"
    ),
    workspace=Depends(get_current_workspace),
    session: AsyncSession = Depends(get_session),
):
    """Export products to Excel or CSV with optional date-range filtering."""
    if format not in _VALID_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"format must be one of: {', '.join(_VALID_FORMATS)}",
        )

    dt_from, dt_to = _parse_date_bounds(date_from, date_to)

    conditions = [Product.workspace_id == workspace.id]
    if dt_from:
        conditions.append(Product.created_at >= dt_from)
    if dt_to:
        conditions.append(Product.created_at <= dt_to)

    result = await session.execute(select(Product).where(*conditions))
    products = result.scalars().all()
    rows = [_product_row(p) for p in products]

    return await _build_export_response(
        rows=rows,
        headers=PRODUCT_HEADERS,
        sheet_title="商品列表",
        filename_base=f"products_{workspace_slug}",
        fmt=format,
    )


@router.get("/orders")
async def export_orders(
    workspace_slug: str,
    format: str = Query("excel", description="Output format: excel or csv"),
    date_from: Optional[str] = Query(
        None, description="Filter orders created on or after this date (YYYY-MM-DD)"
    ),
    date_to: Optional[str] = Query(
        None, description="Filter orders created on or before this date (YYYY-MM-DD)"
    ),
    workspace=Depends(get_current_workspace),
    session: AsyncSession = Depends(get_session),
):
    """Export orders to Excel or CSV with optional date-range filtering."""
    if format not in _VALID_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"format must be one of: {', '.join(_VALID_FORMATS)}",
        )

    dt_from, dt_to = _parse_date_bounds(date_from, date_to)

    conditions = [Order.workspace_id == workspace.id]
    if dt_from:
        conditions.append(Order.created_at >= dt_from)
    if dt_to:
        conditions.append(Order.created_at <= dt_to)

    result = await session.execute(select(Order).where(*conditions))
    orders = result.scalars().all()
    rows = [_order_row(o) for o in orders]

    return await _build_export_response(
        rows=rows,
        headers=ORDER_HEADERS,
        sheet_title="订单列表",
        filename_base=f"orders_{workspace_slug}",
        fmt=format,
    )
